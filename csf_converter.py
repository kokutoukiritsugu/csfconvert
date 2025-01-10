# -*- coding: UTF-8 -*-
import sys
import struct
import openpyxl

def csf_to_xlsx(csf_filename, xlsx_filename):
    encoding_outfile = 'utf-8'
    encoding_CSF_str = 'utf-16le'

    sHead_CSF_Converter = 'head_csf_converter'

    dCSF = {}

    with open(csf_filename, mode='rb') as fCSF:
        bHeadCSF, i1, iStringCount1, iStringCount2, i2, i3 = struct.unpack("<4sLLLLL", fCSF.read(24))
        sHeadCSF = bHeadCSF.decode("ascii")

        dCSF[sHead_CSF_Converter] = {
            "HeadCSF": sHeadCSF,
            "UnknowVaule1": i1,
            "UnknowVaule2": i2,
            "UnknowVaule3": i3
        }

        n = 0
        while not n == iStringCount1:
            bHeadStr, i4, iStringLength = struct.unpack("<4sLL", fCSF.read(12))
            bStringKey, = struct.unpack("<%ds" % iStringLength, fCSF.read(iStringLength))
            sStringKey = bStringKey.decode("ascii")
            dCSF[sStringKey] = {
                'HeadStr': bHeadStr.decode("ascii"),
                'UnknowVaule4': i4,
                'StringType': '',
                'String1Value': '',
                'String2Value': ''
            }

            bStringType, = struct.unpack("<4s", fCSF.read(4))
            dCSF[sStringKey]["StringType"] = bStringType.decode("ascii")

            if bStringType:
                iString1ValueLengthRaw, = struct.unpack("<L", fCSF.read(4))
                bString1ValueRaw, = struct.unpack("<%ds" % (iString1ValueLengthRaw * 2), fCSF.read(iString1ValueLengthRaw * 2))
                bString1Value = b''
                for b in bString1ValueRaw:
                    bString1Value = bString1Value + (b ^ 0xff).to_bytes(1, "little")
                dCSF[sStringKey]["String1Value"] = bString1Value.decode(encoding_CSF_str)

            if bStringType == b'WRTS':
                iString2ValueLength, = struct.unpack("<L", fCSF.read(4))
                bString2Value, = struct.unpack("<%ds" % iString2ValueLength, fCSF.read(iString2ValueLength))
                dCSF[sStringKey]["String2Value"] = bString2Value.decode("ascii")

            n += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CSF Data"

    ws.append(["head_csf_converter"])
    ws.append(["HeadCSF", "UnknowVaule1", "UnknowVaule2", "UnknowVaule3"])
    ws.append([
        dCSF[sHead_CSF_Converter]["HeadCSF"],
        dCSF[sHead_CSF_Converter]["UnknowVaule1"],
        dCSF[sHead_CSF_Converter]["UnknowVaule2"],
        dCSF[sHead_CSF_Converter]["UnknowVaule3"]
    ])

    ws.append([])

    headers = ['Key', 'HeadStr', 'UnknowVaule4', 'StringType', 'String1Value', 'String2Value']
    ws.append(headers)

    column_widths = [36, 9, 14, 12, 150, 14]
    for i, width in enumerate(column_widths, start=1):
        column_letter = openpyxl.utils.get_column_letter(i)
        ws.column_dimensions[column_letter].width = width

    for col in ws.iter_cols(min_row=5, max_row=5):
        for cell in col:
            cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
            cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)

    for key, value in dCSF.items():
        if key != sHead_CSF_Converter:
            row = [
                key,
                value.get('HeadStr', ''),
                value.get('UnknowVaule4', ''),
                value.get('StringType', ''),
                value.get('String1Value', ''),
                value.get('String2Value', '')
            ]
            ws.append(row)

            for cell in ws[ws.max_row]:
                cell.number_format = openpyxl.styles.numbers.FORMAT_TEXT
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb.save(xlsx_filename)
    print(f"Conversion to XLSX completed: {xlsx_filename}")

def xlsx_to_csf(xlsx_filename, csf_filename):
    encoding_outfile = 'utf-8'
    encoding_CSF_str = 'utf-16le'

    sHead_CSF_Converter = 'head_csf_converter'

    wb = openpyxl.load_workbook(xlsx_filename)
    ws = wb.active

    head_csf_converter_data = []
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        head_csf_converter_data.append(row)

    if head_csf_converter_data[0][0] != sHead_CSF_Converter:
        print("error: head_csf_converter information not found in Excel file!")
        sys.exit()

    head_csf_converter = {
        "HeadCSF": head_csf_converter_data[2][0],
        "UnknowVaule1": head_csf_converter_data[2][1],
        "UnknowVaule2": head_csf_converter_data[2][2],
        "UnknowVaule3": head_csf_converter_data[2][3]
    }

    dCSF = {}
    dCSF[sHead_CSF_Converter] = head_csf_converter

    for row in ws.iter_rows(min_row=6, values_only=True):
        if not row[0]:
            continue

        sStringKey, sHeadStr, i4, sStringType, sString1Value, sString2Value = row

        if sString1Value is None:
            sString1Value = ""
        else:
            sString1Value = str(sString1Value)

        if sString2Value is None:
            sString2Value = ""
        else:
            sString2Value = str(sString2Value)

        try:
            i4 = int(i4)
        except (ValueError, TypeError):
            print(f"error: invalid value for UnknowVaule4 in row {row}")
            sys.exit()

        dCSF[sStringKey] = {
            "HeadStr": sHeadStr,
            "UnknowVaule4": i4,
            "StringType": sStringType,
            "String1Value": sString1Value,
            "String2Value": sString2Value
        }

    with open(csf_filename, 'wb') as fCSF:
        sHeadCSF = dCSF[sHead_CSF_Converter]['HeadCSF']
        i1 = dCSF[sHead_CSF_Converter]['UnknowVaule1']
        i2 = dCSF[sHead_CSF_Converter]['UnknowVaule2']
        i3 = dCSF[sHead_CSF_Converter]['UnknowVaule3']
        iStringCount1 = len(dCSF) - 1
        iStringCount2 = iStringCount1
        p = struct.pack("<4sLLLLL", bytes(sHeadCSF, encoding='ascii'), i1, iStringCount1, iStringCount2, i2, i3)
        fCSF.write(p)

        for sStringKey in dCSF:
            if sStringKey == sHead_CSF_Converter:
                continue

            sHeadStr = dCSF[sStringKey]['HeadStr']
            i4 = dCSF[sStringKey]['UnknowVaule4']
            iStringLength = len(sStringKey)
            p = struct.pack("<4sLL%ds" % iStringLength, bytes(sHeadStr, encoding='ascii'), i4, iStringLength, bytes(sStringKey, encoding='ascii'))
            fCSF.write(p)

            sStringType = dCSF[sStringKey]['StringType']
            sString1Value = dCSF[sStringKey]['String1Value']

            if sStringType:
                if not isinstance(sString1Value, str):
                    sString1Value = str(sString1Value)

                bString1Value = bytes(sString1Value, encoding=encoding_CSF_str)
                bString1ValueRaw = b''
                for b in bString1Value:
                    bString1ValueRaw = bString1ValueRaw + (b ^ 0xff).to_bytes(1, "little")
                iString1ValueLength = len(bString1ValueRaw)
                iString1ValueLengthRaw = iString1ValueLength // 2
                p = struct.pack("<4sL%ds" % iString1ValueLength, bytes(sStringType, encoding='ascii'), iString1ValueLengthRaw, bString1ValueRaw)
                fCSF.write(p)

            if sStringType == "WRTS":
                sString2Value = dCSF[sStringKey]["String2Value"]
                if not isinstance(sString2Value, str):
                    sString2Value = str(sString2Value)

                iString2ValueLength = len(sString2Value)
                p = struct.pack("<L%ds" % iString2ValueLength, iString2ValueLength, bytes(sString2Value, encoding='ascii'))
                fCSF.write(p)

    print(f"Conversion to CSF completed: {csf_filename}")

def main():
    print("Red Alert 2 CSF language file converter")
    print("by kokutoukiritsugu")
    print()

    if len(sys.argv) != 4:
        print("usage:")
        print("  Convert CSF to XLSX: python3 %s csf2xlsx csfFilename xlsxFilename" % sys.argv[0])
        print("  Convert XLSX to CSF: python3 %s xlsx2csf xlsxFilename csfFilename" % sys.argv[0])
        sys.exit()

    mode = sys.argv[1]
    input_file = sys.argv[2]
    output_file = sys.argv[3]

    if mode == "csf2xlsx":
        csf_to_xlsx(input_file, output_file)
    elif mode == "xlsx2csf":
        xlsx_to_csf(input_file, output_file)
    else:
        print("error: invalid mode!")
        sys.exit()

if __name__ == "__main__":
    main()